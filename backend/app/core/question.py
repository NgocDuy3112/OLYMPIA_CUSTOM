import json
import zipfile
import mimetypes
from sqlalchemy import select, text

for _ext, _mime in [
    (".jpg", "image/jpeg"), (".jpeg", "image/jpeg"), (".png", "image/png"),
    (".gif", "image/gif"), (".webp", "image/webp"), (".svg", "image/svg+xml"),
    (".mp3", "audio/mpeg"), (".ogg", "audio/ogg"), (".wav", "audio/wav"),
    (".aac", "audio/aac"), (".m4a", "audio/mp4"),
    (".mp4", "video/mp4"), (".webm", "video/webm"), (".mov", "video/quicktime"),
]:
    mimetypes.add_type(_mime, _ext)
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.exc import IntegrityError

from logger import global_logger
from models.question import Question
from models.match import Match
from schemas.question import *
from configs import *
from fastapi import HTTPException, UploadFile
from openpyxl import load_workbook
from io import BytesIO

QUESTION_SHEET_NAMES = ['KHOI_DONG', 'GIAI_MA', 'BUT_PHA', 'VE_DICH']
_MATCH_PATTERN = AppSettings().MATCH_PATTERN


def _normalize_media_url(raw, match_code: str) -> str | None:
    if raw is None:
        return None
    v = str(raw).strip()
    if not v or v == "None":
        return None
    if v.startswith("http://") or v.startswith("https://") or v.startswith(_MATCH_PATTERN):
        return v
    return f"{match_code}/{v}"


async def post_questions_from_excel_to_db(
    match_code: str,
    file: UploadFile,
    session: AsyncSession,
    overwrite: bool = False
) -> BaseResponse:
    global_logger.debug(f"POST request received to inject questions from Excel with match code: {match_code}, overwrite={overwrite}.")
    try:
        match_id = await session.scalar(select(Match.id).where(Match.match_code == match_code))
        if match_id is None:
            log_message = f"No match found with match_code={match_code}."
            global_logger.warning(log_message)
            raise HTTPException(status_code=404, detail=log_message)

        if overwrite:
            for dep in ("answers", "records"):
                await session.execute(
                    text(f"DELETE FROM {dep} WHERE match_id = :match_id"),
                    {"match_id": match_id},
                )
            await session.execute(
                text("DELETE FROM questions WHERE match_id = :match_id"),
                {"match_id": match_id},
            )
            await session.commit()
            global_logger.info(f"Deleted existing questions for match_code={match_code} in overwrite mode.")

        content = await file.read()
        wb = load_workbook(BytesIO(content), data_only=True)

        for sheet_name in QUESTION_SHEET_NAMES:
            if sheet_name not in wb.sheetnames:
                global_logger.debug(f"Sheet '{sheet_name}' not found in uploaded workbook; skipping.")
                continue
            ws = wb[sheet_name]

            rows = []
            for row in ws.iter_rows(min_row=2, max_col=5, values_only=True):
                if not row:
                    continue
                if len(row) <= 1 or not any((cell is not None and str(cell).strip()) for cell in row[1:]):
                    continue
                qcode = row[0]
                content_cell = row[1]
                answer_cell = row[2] if len(row) > 2 else None
                explanation_cell = row[3] if len(row) > 3 else None
                media_cell = row[4] if len(row) > 4 else None

                if not qcode or not content_cell or not answer_cell:
                    global_logger.warning(f"Skipping invalid row in sheet '{sheet_name}': {row}")
                    continue

                rows.append((qcode, content_cell, answer_cell, explanation_cell, media_cell))

            question_objects = []
            for r in rows:
                try:
                    question_objects.append(Question(
                            question_code=str(r[0]).strip(),
                            content=str(r[1]),
                            answer=str(r[2]),
                            explanation=str(r[3]) if r[3] is not None else None,
                            media_url=_normalize_media_url(r[4], match_code),
                            match_id=match_id
                        ))
                except Exception as e:
                    global_logger.error(f"Failed constructing Question object for row {r}: {e}", exc_info=True)

            if question_objects:
                session.add_all(question_objects)
                global_logger.debug(f"Added {len(question_objects)} questions from sheet '{sheet_name}' to session.")

        await session.commit()
        log_message = f"Questions injected successfully from Excel for match_code={match_code}."
        global_logger.info(log_message)
        return BaseResponse(
            status='success',
            message=log_message
        )
    except IntegrityError:
        await session.rollback()
        log_message = f"Question in match_code={match_code} already exists."
        global_logger.warning(log_message)
        raise HTTPException(status_code=400, detail=log_message)
    except HTTPException:
        raise
    except Exception as e:
        await session.rollback()
        log_message = f"An unexpected error occurred while injecting questions from Excel with match_code={match_code}: {e}"
        global_logger.exception(log_message)
        raise HTTPException(status_code=500, detail=log_message)



async def post_question_to_db(
    request: QuestionPostRequest, 
    session: AsyncSession
) -> BaseResponse:
    global_logger.debug(f"POST request received to add question with code: {request.question_code}.")
    try:
        match_id = await session.scalar(select(Match.id).where(Match.match_code == request.match_code, Match.is_deleted == False))
        if match_id is None:
            log_message = f"No match found with match_code={request.match_code}."
            global_logger.warning(log_message)
            raise HTTPException(status_code=404, detail=log_message)
        opts_value = None
        if request.options is not None:
            if isinstance(request.options, list):
                try:
                    opts_value = json.dumps(request.options, ensure_ascii=False)
                except Exception:
                    opts_value = None
            elif isinstance(request.options, str):
                opts_value = request.options

        question = Question(
            question_code=request.question_code,
            content=request.content,
            answer=request.answer,
            explanation=request.explanation,
            media_url=request.media_url,
            options=opts_value,
            match_id=match_id
        )
        session.add(question)
        await session.commit()
        log_message = f"Question with question_code={request.question_code} added successfully to the database."
        global_logger.info(log_message)
        return BaseResponse(
            status='success',
            message=log_message
        )
    except IntegrityError:
        await session.rollback()
        log_message = f"Question with question_code={request.question_code} already exists."
        global_logger.warning(log_message)
        raise HTTPException(status_code=400, detail=log_message)
    except HTTPException:
        raise
    except Exception:
        await session.rollback()
        log_message = f"An unexpected error occurred while adding question with question_code={request.question_code}."
        global_logger.exception(log_message)
        raise HTTPException(status_code=500, detail=log_message)


async def get_question_from_request_from_db(
    match_code: str,
    question_code: str | None, 
    session: AsyncSession
) -> BaseResponse:
    global_logger.debug(f"GET request received to fetch question with code: {question_code}.")
    question_data = []
    try:
        if question_code is not None:
            query = select(Question).where(
                Question.question_code == question_code,
                Question.is_deleted == False,
                Question.match_id == select(Match.id).where(
                    Match.match_code == match_code)
                .scalar_subquery()
            )
            result = await session.scalar(query)
            question = result
            if question is None:
                log_message = f"No question found with question_code={question_code}."
                global_logger.warning(log_message)
                raise HTTPException(status_code=400, detail=log_message)
            parsed_options = None
            if question.options is not None:
                try:
                    parsed = json.loads(question.options)
                    if isinstance(parsed, list):
                        parsed_options = parsed
                    else:
                        parsed_options = question.options
                except Exception:
                    parsed_options = question.options

            question_data = {
                'question_code': question.question_code,
                'content': question.content,
                'answer': question.answer,
                'explanation': question.explanation,
                'media_url': question.media_url,
                'options': parsed_options,
            }
        else:
            query = select(Question).where(
                Question.is_deleted == False,
                Question.match_id == select(Match.id).where(
                    Match.match_code == match_code)
                .scalar_subquery()
            )
            result = await session.scalars(query)
            questions = result.all()
            question_data = []
            for q in questions:
                parsed_options = None
                if q.options is not None:
                    try:
                        parsed = json.loads(q.options)
                        parsed_options = parsed if isinstance(parsed, list) else q.options
                    except Exception:
                        parsed_options = q.options
                question_data.append({
                    'question_code': q.question_code,
                    'content': q.content,
                    'answer': q.answer,
                    'explanation': q.explanation,
                    'media_url': q.media_url,
                    'options': parsed_options,
                })
        log_message = f"Fetched {len(question_data)} questions from the database with question_code={question_code}."
        global_logger.debug(log_message)
        return BaseResponse(
            status='success',
            message=log_message,
            data=question_data
        )
    except HTTPException:
        raise
    except Exception as e:
        log_message = f"An unexpected error occurred while fetching question with question_code={question_code}: {str(e)}"
        global_logger.exception(log_message)
        raise HTTPException(status_code=500, detail=log_message)


async def delete_question_from_db(match_code: str, question_code: str, session: AsyncSession) -> BaseResponse:
    global_logger.debug(f"Soft deleting question with question_code={question_code} in match_code={match_code} from database.")
    try:
        match_id = await session.scalar(select(Match.id).where(Match.match_code == match_code, Match.is_deleted == False))
        if match_id is None:
            log_message = f"No active match found with match_code={match_code}."
            global_logger.warning(log_message)
            raise HTTPException(status_code=404, detail=log_message)

        query = select(Question).where(
            Question.question_code == question_code,
            Question.match_id == match_id,
            Question.is_deleted == False
        )
        result = await session.execute(query)
        question = result.scalars().one_or_none()

        if question is None:
            log_message = f"No active question found with question_code={question_code} in match_code={match_code} to delete."
            global_logger.warning(log_message)
            raise HTTPException(status_code=404, detail=log_message)

        question.is_deleted = True
        await session.commit()

        log_message = f"Question with question_code={question_code} in match_code={match_code} has been soft deleted successfully."
        global_logger.info(log_message)
        return BaseResponse(
            status='success',
            message=log_message
        )
    except HTTPException:
        raise
    except Exception:
        await session.rollback()
        log_message = f"An unexpected error occurred while deleting question with question_code={question_code}."
        global_logger.exception(log_message)
        raise HTTPException(status_code=500, detail=log_message)


async def patch_question_to_db(
    match_code: str,
    question_code: str,
    request: QuestionUpdateRequest,
    session: AsyncSession,
) -> BaseResponse:
    global_logger.info(f"PATCH request received to update question_code={question_code} in match_code={match_code}.")
    try:
        match_id = await session.scalar(
            select(Match.id).where(Match.match_code == match_code, Match.is_deleted == False)
        )
        if match_id is None:
            log_message = f"No active match found with match_code={match_code}."
            global_logger.warning(log_message)
            raise HTTPException(status_code=404, detail=log_message)

        result = await session.execute(
            select(Question).where(
                Question.question_code == question_code,
                Question.match_id == match_id,
                Question.is_deleted == False,
            )
        )
        question = result.scalar_one_or_none()
        if question is None:
            log_message = f"No active question found with question_code={question_code} in match_code={match_code}."
            global_logger.warning(log_message)
            raise HTTPException(status_code=404, detail=log_message)

        if request.content is not None:
            question.content = request.content
        if request.answer is not None:
            question.answer = request.answer
        if request.explanation is not None:
            question.explanation = request.explanation
        if request.media_url is not None:
            question.media_url = request.media_url
        if request.options is not None:
            if isinstance(request.options, list):
                try:
                    question.options = json.dumps(request.options, ensure_ascii=False)
                except Exception:
                    question.options = None
            else:
                question.options = request.options

        await session.commit()
        log_message = f"Question updated successfully for question_code={question_code} in match_code={match_code}."
        global_logger.info(log_message)
        return BaseResponse(status='success', message=log_message)
    except HTTPException:
        raise
    except Exception:
        await session.rollback()
        log_message = f"An unexpected error occurred while updating question with question_code={question_code}."
        global_logger.exception(log_message)
        raise HTTPException(status_code=500, detail=log_message)


_MEDIA_MIME_TYPES: frozenset[str] = frozenset({
    "image/jpeg", "image/png", "image/gif", "image/webp", "image/svg+xml",
    "audio/mpeg", "audio/ogg", "audio/wav", "audio/webm", "audio/mp4", "audio/aac",
    "video/mp4", "video/webm", "video/ogg", "video/quicktime",
})


async def post_questions_from_zip_to_db(
    file: UploadFile,
    s3_client,
    bucket: str,
    session: AsyncSession,
    overwrite: bool = False,
) -> BaseResponse:
    filename = (file.filename or "").strip()
    stem = filename.rsplit(".", 1)[0] if "." in filename else filename
    match_code = stem

    global_logger.info(f"ZIP import started: file='{filename}', match_code='{match_code}'")

    if not match_code.startswith(_MATCH_PATTERN):
        raise HTTPException(
            status_code=400,
            detail=f"Tên file ZIP phải bắt đầu bằng '{_MATCH_PATTERN}', nhận được: '{filename}'.",
        )

    raw = await file.read()
    try:
        zf = zipfile.ZipFile(BytesIO(raw))
    except zipfile.BadZipFile:
        raise HTTPException(status_code=400, detail="File không phải định dạng ZIP hợp lệ.")

    names = zf.namelist()

    excel_name = f"{match_code}.xlsx"
    excel_entry = next(
        (n for n in names if n.split("/")[-1] == excel_name and not n.startswith("__")),
        None,
    )
    if excel_entry is None:
        raise HTTPException(
            status_code=400,
            detail=f"Không tìm thấy '{excel_name}' trong ZIP.",
        )

    media_ok: list[str] = []
    media_fail: list[str] = []

    for entry in names:
        basename = entry.split("/")[-1]
        if not basename or basename == excel_name or entry.startswith("__"):
            continue

        mime, _ = mimetypes.guess_type(basename)
        if not mime or mime not in _MEDIA_MIME_TYPES:
            global_logger.warning(f"ZIP: bỏ qua '{entry}' (MIME không hợp lệ: {mime!r})")
            continue

        key = f"{match_code}/{basename}"
        try:
            data = zf.read(entry)
            await s3_client.put_object(
                Bucket=bucket, Key=key, Body=data,
                ContentType=mime, ContentLength=len(data),
            )
            media_ok.append(basename)
            global_logger.info(f"ZIP: S3 upload ok key='{key}' ({len(data)} bytes)")
        except Exception as exc:
            media_fail.append(basename)
            global_logger.warning(
                f"ZIP: S3 upload failed for key='{key}' ({len(data)} bytes): {exc}",
                exc_info=True,
            )

    excel_bytes = zf.read(excel_entry)
    excel_file = UploadFile(filename=excel_name, file=BytesIO(excel_bytes))

    result = await post_questions_from_excel_to_db(match_code, excel_file, session, overwrite=overwrite)

    summary_parts = [result.message]
    if media_ok:
        summary_parts.append(f"Media uploaded: {len(media_ok)} file(s).")
    if media_fail:
        summary_parts.append(f"Media skipped (lỗi): {', '.join(media_fail)}.")

    return BaseResponse(
        status="error" if media_fail else "success",
        message=" | ".join(summary_parts),
    )
